import discord
from discord.ext import commands, tasks
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from googleapiclient.discovery import build
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.oauth2.credentials import Credentials
import os
import requests
from datetime import datetime, timedelta


load_dotenv()

TOKEN = os.getenv("DISCORD_TOKEN")
GOOGLE_CREDENTIALS = 'credentials.json'
NOTIFICATION_CHANNEL = int(os.getenv("NOTIFICATION_CHANNEL"))
folder_id = "xxxxxxxxxxxxxxxx"
user_images = {}


intents = discord.Intents.all()
intents.message_content = True
intents.members = True

def get_dollar_exchange_rate():
    """Obtém a cotação atual do dólar em relação ao BRL."""
    API_URL = "https://api.exchangerate-api.com/v4/latest/USD"  
    try:
        response = requests.get(API_URL)
        response.raise_for_status()
        data = response.json()
        return data["rates"].get("BRL", None)  
    except Exception as e:
        print(f"❌ Erro ao buscar cotação do dólar: {e}")
        return None
    
# Inicializar o bot
bot = commands.Bot(
    command_prefix="!",
    intents=intents,
    description="Bot para gerenciamento de pagamentos",
    help_command=commands.DefaultHelpCommand()
)

# Classe para gerenciamento de pagamentos
import io
from googleapiclient.http import MediaIoBaseDownload

class PaymentManager:
    def __init__(self, filename="payments.xlsx"):
        self.filename = filename
        self.payment_day = 15
        self.data = {"members": {}, "account_links": {}, "auto_paid_members": []}
        if not os.path.exists(self.filename):
            self.create_empty_file()
        self.load_data()

    def link_account(self, secondary_id, main_id):
        """Vincula uma conta secundária a uma principal"""
        self.data["account_links"][str(secondary_id)] = str(main_id)
        self.save_data()

    def get_main_account(self, user_id):
        """Retorna a conta principal associada a um usuário, ou o próprio ID se não houver vínculo"""
        return self.data["account_links"].get(str(user_id), str(user_id))

    def set_payment_day(self, day: int):
        """Define o dia de pagamento e salva no arquivo"""
        if day < 1 or day > 31:
            print("❌ O dia de pagamento deve ser entre 1 e 31.")
            return
        self.payment_day = day
        self.save_data()
        print(f"✅ Dia de pagamento atualizado para {day}.")



    def create_empty_file(self):
        """Cria um arquivo Excel vazio"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["UserID", "Username", "Payments"])  
        workbook.save(self.filename)
        print(f"✅ Arquivo vazio criado localmente: {self.filename}")


    def load_data(self):
        """Carrega dados do arquivo Excel, incluindo o dia de pagamento."""
        if os.path.exists(self.filename):
            workbook = load_workbook(self.filename, data_only=True)
            sheet = workbook.active

            self.data["members"] = {
                str(row[0]): {"username": row[1], "payments": eval(row[2])}
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] is not None
            }

            if "Config" in workbook.sheetnames:
                config_sheet = workbook["Config"]
                self.payment_day = int(config_sheet.cell(row=2, column=1).value) 
            else:
                config_sheet = workbook.create_sheet("Config")
                config_sheet.append(["PaymentDay"])
                config_sheet.append([int(self.payment_day)])
                workbook.save(self.filename)

            if "AutoPaid" in workbook.sheetnames:
                auto_paid_sheet = workbook["AutoPaid"]
                self.data["auto_paid_members"] = [
                    str(row[0]) for row in auto_paid_sheet.iter_rows(min_row=2, values_only=True) if row[0]
                ]
            else:
                self.data["auto_paid_members"] = []

            if "AccountLinks" in workbook.sheetnames:
                links_sheet = workbook["AccountLinks"]
                self.data["account_links"] = {
                    str(row[0]): str(row[1])
                    for row in links_sheet.iter_rows(min_row=2, values_only=True)
                    if row[0] and row[1]
                }
            else:
                self.data["account_links"] = {}
        else:
            self.save_data()




    def load_from_google_drive(self):
        """Baixa os dados mais recentes do Google Drive ou cria o arquivo se não existir"""
        SCOPES = ['https://www.googleapis.com/auth/drive']
        creds = Credentials.from_authorized_user_file(GOOGLE_CREDENTIALS, SCOPES)
        service = build('drive', 'v3', credentials=creds)

       
        query = f"name='{os.path.basename(self.filename)}'"
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        files = response.get('files', [])

        if files:
            file_id = files[0]['id']
            request = service.files().get_media(fileId=file_id)
            with open(self.filename, 'wb') as file:
                downloader = MediaIoBaseDownload(file, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
            print(f"✅ Arquivo carregado com sucesso do Google Drive: {self.filename}")
        else:
            print("❌ Arquivo não encontrado no Google Drive. Criando novo arquivo...")
            self.create_file_in_drive(service)

        self.load_data()

    def create_file_in_drive(self, service):
        """Cria o arquivo no Google Drive"""
        file_metadata = {'name': os.path.basename(self.filename)}
        media = MediaFileUpload(self.filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Criação do arquivo no Google Drive
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_id = file.get('id')
        print(f"✅ Arquivo criado com sucesso no Google Drive. Acesse em: https://drive.google.com/file/d/{file_id}")


    def save_data(self):
        """Salva os dados no arquivo Excel, incluindo o dia de pagamento."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Payments"
        sheet.append(["UserID", "Username", "Payments"])  # Cabeçalhos

        # Salva todos os membros com seus registros de pagamentos
        for user_id, info in self.data["members"].items():
            sheet.append([user_id, info["username"], str(info["payments"])])

        # Salva o dia de pagamento em uma aba separada
        config_sheet = workbook.create_sheet("Config")
        config_sheet.append(["PaymentDay"])
        config_sheet.append([self.payment_day])

        if self.data["auto_paid_members"]:
            auto_paid_sheet = workbook.create_sheet("AutoPaid")
            auto_paid_sheet.append(["UserID"])  # Cabeçalho
            for user_id in self.data["auto_paid_members"]:
                auto_paid_sheet.append([user_id])

        if self.data["account_links"]:
            links_sheet = workbook.create_sheet("AccountLinks")
            links_sheet.append(["SecondaryID", "MainID"])  # Cabeçalhos
            for secondary_id, main_id in self.data["account_links"].items():
                links_sheet.append([secondary_id, main_id])

        workbook.save(self.filename)
        print(f"✅ Dados salvos com sucesso no arquivo {self.filename}.")


    def upload_image_to_google_drive(self, image_path, image_name):
        """
        Faz upload de uma imagem para o Google Drive na pasta especificada e exclui o arquivo local após o upload.
        """
        SCOPES = ['https://www.googleapis.com/auth/drive']
        creds = Credentials.from_authorized_user_file(GOOGLE_CREDENTIALS, SCOPES)
        service = build('drive', 'v3', credentials=creds)

        file_metadata = {
            'name': image_name,  # Nome do arquivo no formato username - datapagamento
            'parents': [folder_id]  # Especifica a pasta onde o arquivo será armazenado
        }
        media = MediaFileUpload(image_path, mimetype='image/jpeg')  # Assumindo imagens no formato JPEG

        try:
            # Upload do arquivo para o Google Drive
            file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
            print(f"✅ Imagem enviada para o Google Drive: {image_name}")

            # Exclui o arquivo local após o upload bem-sucedido
            os.remove(image_path)
            print(f"🗑️ Arquivo local {image_path} excluído com sucesso.")
        except Exception as e:
            print(f"❌ Falha ao enviar ou excluir o arquivo {image_path}: {e}")

    def upload_to_google_drive(self):
        """
        Faz upload do arquivo Excel para o Google Drive, substituindo o arquivo existente.
        Também exclui arquivos locais temporários após o upload bem-sucedido.
        """
        SCOPES = ['https://www.googleapis.com/auth/drive']
        creds = Credentials.from_authorized_user_file(GOOGLE_CREDENTIALS, SCOPES)
        service = build('drive', 'v3', credentials=creds)

        # Verifica se o arquivo Excel já existe na pasta especificada
        query = f"name='{os.path.basename(self.filename)}' and '{folder_id}' in parents"
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        files = response.get('files', [])
        print(f"Debug: Arquivos encontrados: {files}")

        try:
            if files:
                # Atualiza o arquivo existente
                file_id = files[0]['id']
                file_metadata = {'name': os.path.basename(self.filename)}
                media = MediaFileUpload(self.filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                updated_file = service.files().update(fileId=file_id, body=file_metadata, media_body=media).execute()
                print(f"✅ Arquivo atualizado no Google Drive: {os.path.basename(self.filename)}")
            else:
                # Cria um novo arquivo na pasta especificada
                file_metadata = {
                    'name': os.path.basename(self.filename),
                    'parents': [folder_id]  # Especifica a pasta onde o arquivo será armazenado
                }
                media = MediaFileUpload(self.filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
                print(f"✅ Arquivo criado no Google Drive na pasta especificada: {os.path.basename(self.filename)}")

            # Listar todos os arquivos na pasta especificada para depuração
            response = service.files().list(
                q=f"'{folder_id}' in parents",
                spaces='drive',
                fields='files(id, name)'
            ).execute()
            print(f"Arquivos na pasta do Google Drive: {response.get('files', [])}")

            # Exclui o arquivo local após o upload (apenas se o upload foi bem-sucedido)
            try:
                os.remove(self.filename)
                print(f"🗑️ Arquivo local {self.filename} excluído com sucesso.")
            except Exception as e:
                print(f"❌ Falha ao excluir o arquivo local {self.filename}: {e}")

        except Exception as e:
            print(f"❌ Erro durante o upload para o Google Drive: {e}")



    def register_payment(self, user_id, username):
        """Registra o pagamento do usuário, considerando o vencimento"""
        today = datetime.now()
        user_id_str = str(user_id)

        # Determinar o mês de registro do pagamento
        if today.day > self.payment_day:
            # Caso a data de vencimento já tenha passado, registre no próximo mês
            first_day_next_month = today.replace(day=28) + timedelta(days=4)
            payment_month = first_day_next_month.replace(day=1).strftime("%Y-%m")
        else:
            # Caso contrário, registre no mês atual
            payment_month = today.strftime("%Y-%m")

        # Garantir que o usuário está registrado
        if user_id_str not in self.data["members"]:
            self.data["members"][user_id_str] = {"username": username, "payments": {}}

        # Registrar o pagamento
        self.data["members"][user_id_str]["payments"][payment_month] = True
        self.save_data()


    def get_payment_status(self):
        """Retorna o status de pagamento de todos os membros"""
        today = datetime.now()
        current_month = today.strftime("%Y-%m")
        
        # Verifica se o dia atual é depois do vencimento
        if today.day > self.payment_day:
            payment_month = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
        else:
            payment_month = current_month
        
        return [
            {"username": info["username"], "paid": info["payments"].get(payment_month, False)}
            for info in self.data["members"].values()
        ]

    def register_initial_payment(self, user_id, username):
        """Registra um novo membro no sistema com status de pagamento pendente."""
        current_month = datetime.now().strftime("%Y-%m")
        user_id_str = str(user_id)  # Garantir que o ID está como string

        if user_id_str not in self.data["members"]:
            # Adiciona o usuário ao sistema
            self.data["members"][user_id_str] = {"username": username, "payments": {}}

        # Garante que o status do mês atual seja 'pendente'
        if current_month not in self.data["members"][user_id_str]["payments"]:
            self.data["members"][user_id_str]["payments"][current_month] = False  # Pendente

        # Salva no arquivo Excel
        self.save_data()

    def has_paid(self, user_id):
        """Verifica se o usuário já pagou no mês atual"""
        current_month = datetime.now().strftime("%Y-%m")
        return user_id in self.data["members"] and current_month in self.data["members"][user_id]["payments"]

    def unregister_payment(self, user_id):
        """Remove o pagamento do usuário considerando o vencimento"""
        today = datetime.now()
        user_id_str = str(user_id)

        # Determinar o mês relevante para remoção
        if today.day > self.payment_day:
            # Após o vencimento, remove do próximo mês
            first_day_next_month = today.replace(day=28) + timedelta(days=4)
            payment_month = first_day_next_month.replace(day=1).strftime("%Y-%m")
        else:
            # Antes ou no vencimento, remove do mês atual
            payment_month = today.strftime("%Y-%m")

        # Verificar e remover o pagamento
        if user_id_str in self.data["members"]:
            payments = self.data["members"][user_id_str].get("payments", {})
            
            if payment_month in payments:
                payments[payment_month] = False
                self.save_data()  
                self.upload_to_google_drive()
                print(f"✅ Pagamento removido para o usuário {user_id_str} no mês {payment_month}.")
                return True
            else:
                print(f"❌ Nenhum pagamento encontrado para o usuário {user_id_str} no mês {payment_month}.")
                return False
        else:
            print(f"❌ Usuário {user_id_str} não está registrado.")
            return False

payment_manager = PaymentManager()

@bot.event
async def on_ready():
    print(f"✅ Bot conectado como {bot.user}")
    guild = bot.guilds[0]
    payment_manager.load_from_google_drive()
    for member in guild.members:
        if not member.bot:
            user_id = str(member.id)
            if user_id in payment_manager.data["auto_paid_members"]:
                payment_manager.register_payment(user_id, member.name)
    check_payments.start()

@bot.event
async def on_message(message):
    if message.attachments:
        for attachment in message.attachments:
            if attachment.filename.lower().endswith(('jpg', 'jpeg', 'png', 'pdf')):
                # Baixa o arquivo e salva localmente
                save_path = f"temp_{message.author.id}_{attachment.filename}"
                await attachment.save(save_path)
                user_images[message.author.id] = save_path
                await message.channel.send(f"📸 Comprovante recebido! Agora você pode usar o comando `!paguei` para finalizar o registro.")
    await bot.process_commands(message)


@bot.command(name='paguei', help='Registra seu pagamento para o mês atual e envia o comprovante')
async def register_payment(ctx):
    """Registra o pagamento do membro, sincroniza com o Google Drive e faz upload do comprovante"""
    user_id = payment_manager.get_main_account(ctx.author.id)
    payment_manager.register_payment(user_id, ctx.author.name)

    # Verifica se o usuário enviou um comprovante de pagamento
    if ctx.author.id in user_images:
        image_path = user_images.pop(ctx.author.id)  # Remove a entrada após usar
        date_str = datetime.now().strftime("%d-%m-%Y")  # Data do pagamento no formato desejado
        image_name = f"{ctx.author.name} - {date_str}.jpg"  # Formata o nome como username - datapagamento
        payment_manager.upload_image_to_google_drive(image_path, image_name)
        await ctx.send(f"✅ Pagamento registrado com sucesso e comprovante enviado para {ctx.author.name}!")
    else:
        await ctx.send(f"✅ Pagamento registrado com sucesso para {ctx.author.name}, mas nenhum comprovante foi encontrado.")

    # Sincroniza os dados com o Google Drive
    payment_manager.upload_to_google_drive()


@bot.command(name='naopaguei', help='Remove seu registro de pagamento do mês atual ou próximo, dependendo do vencimento')
async def unregister_payment(ctx):
    """Desfaz o pagamento do membro"""
    user_id = payment_manager.get_main_account(ctx.author.id)
    if payment_manager.unregister_payment(user_id):
        await ctx.send(f"🔄 {ctx.author.name}, seu pagamento foi removido com sucesso.")
    else:
        await ctx.send(f"❌ {ctx.author.name}, você não tem um pagamento registrado para o período atual ou próximo.")


@bot.command(name='pago', help='Permite ao administrador marcar outro membro como pago para um mês específico.')
@commands.has_permissions(administrator=True)
async def mark_as_paid(ctx, member: discord.Member, month: str = None):
    """
    Permite que um administrador registre o pagamento de outro membro para um mês específico.
    Se o mês não for especificado, usa o mês atual ou próximo.
    """
    user_id = payment_manager.get_main_account(member.id)
    username = member.name

    # Determinar o mês para registrar o pagamento
    if month:
        try:
            # Validar e formatar o mês fornecido
            payment_month = datetime.strptime(month, "%Y-%m").strftime("%Y-%m")
        except ValueError:
            await ctx.send("❌ Formato de mês inválido. Use o formato: `YYYY-MM` (ex: 2025-01).")
            return
    else:
        # Usar a lógica padrão para determinar o mês
        today = datetime.now()
        if today.day > payment_manager.payment_day:
            # Após o vencimento, registrar no próximo mês
            first_day_next_month = today.replace(day=28) + timedelta(days=4)
            payment_month = first_day_next_month.replace(day=1).strftime("%Y-%m")
        else:
            # Antes ou no vencimento, registrar no mês atual
            payment_month = today.strftime("%Y-%m")

    # Garantir que o usuário está registrado
    if user_id not in payment_manager.data["members"]:
        payment_manager.data["members"][user_id] = {"username": username, "payments": {}}

    # Registrar o pagamento para o mês correto
    payment_manager.data["members"][user_id]["payments"][payment_month] = True

    # Salvar e sincronizar os dados
    payment_manager.save_data()
    payment_manager.upload_to_google_drive()

    await ctx.send(f"✅ Pagamento registrado para {member.mention} no mês {payment_month}.")


@bot.command(name='status', help='Mostra o status dos pagamentos de todos os membros')
async def payment_status(ctx):
    """Exibe o status de pagamento de todos os membros"""
    today = datetime.now()
    if today.day > payment_manager.payment_day:
        # Após o dia do vencimento, mova para o próximo mês
        first_day_next_month = today.replace(day=28) + timedelta(days=4)
        relevant_month = first_day_next_month.replace(day=1).strftime("%Y-%m")
    else:
        # Antes ou no dia do vencimento, use o mês atual
        relevant_month = today.strftime("%Y-%m")
    
    status = [
        {"username": info["username"], "paid": info["payments"].get(relevant_month, False)}
        for info in payment_manager.data["members"].values()
    ]
    
    message = f"📊 **Status dos Pagamentos de {relevant_month}**\n\n"
    for member in status:
        emoji = "✅" if member["paid"] else "❌"
        message += f"{emoji} {member['username']} {'(Pendente)' if not member['paid'] else ''}\n"
    await ctx.send(message)


@bot.command(name='preco', help='Exibe o preço atual da assinatura em BRL e por pessoa')
async def show_price(ctx):
    """Calcula e exibe o preço da assinatura com base na cotação do dólar, considerando o mês correto"""
    exchange_rate = get_dollar_exchange_rate()
    if exchange_rate is None:
        await ctx.send("❌ Não foi possível obter a cotação atual do dólar. Tente novamente mais tarde.")
        return

    total_price = 20 * exchange_rate
    price_per_person = total_price / 5

    today = datetime.now()
    if today.day > payment_manager.payment_day:
        # Após o vencimento, defina a data de vencimento do próximo mês
        first_day_next_month = today.replace(day=28) + timedelta(days=4)
        due_date = first_day_next_month.replace(day=payment_manager.payment_day)
    else:
        # Antes ou no vencimento, mantenha a data atual
        due_date = today.replace(day=payment_manager.payment_day)

    due_date_str = due_date.strftime('%d/%m/%Y')
    pix_key = "146.980.396-81"  

    await ctx.send(
        f"💵 Cotação atual do dólar: **R${exchange_rate:.2f}**.\n"
        f"💵 O preço total da assinatura de $20 é **R${total_price:.2f}**.\n"
        f"👥 Dividido por 5 pessoas, cada uma deve pagar **R${price_per_person:.2f}**.\n"
        f"📅 A data de vencimento para o pagamento é **{due_date_str}**.\n"
        f"💳 A chave PIX para pagamento é: **{pix_key}**."
    )


@tasks.loop(hours=24)
async def check_payments():
    """Tarefa automática para verificar pagamentos e notificar membros pendentes."""
    channel = bot.get_channel(NOTIFICATION_CHANNEL)
    if not channel:
        print("❌ Canal de notificações não encontrado.")
        return

    today = datetime.now()
    payment_date = datetime(today.year, today.month, int(payment_manager.payment_day))
    days_until_payment = (payment_date - today).days

    # Processa pagamentos automáticos
    for user_id in payment_manager.data["auto_paid_members"]:
        if user_id in payment_manager.data["members"]:
            payment_manager.register_payment(user_id, payment_manager.data["members"][user_id]["username"])

    if days_until_payment in [2, 0]:
        # Carrega dados do Google Drive antes de verificar o status
        payment_manager.load_from_google_drive()
        status = payment_manager.get_payment_status()

        # Lista de membros pendentes
        pending_members = [
            f"<@{user_id}>" for user_id, info in payment_manager.data["members"].items()
            if not info["payments"].get(datetime.now().strftime("%Y-%m"), False)
        ]

        # Cotação do dólar
        exchange_rate = get_dollar_exchange_rate()
        if exchange_rate is not None:
            total_price = 20 * exchange_rate
            price_per_person = total_price / 5
            price_message = (
                f"\n💵 O preço total da assinatura de $20 é **R${total_price:.2f}**."
                f"\n👥 Dividido por 5 pessoas, cada uma deve pagar **R${price_per_person:.2f}**."
            )
        else:
            price_message = "\n❌ Não foi possível obter a cotação do dólar no momento."

        message = "🔔 **Lembrete de Pagamento**\n"
        if days_until_payment == 2:
            message += "Faltam 2 dias para o pagamento!\n\n"
        else:
            message += "O pagamento vence hoje!\n\n"

        if pending_members:
            message += f"📋 **Membros pendentes:**\n{', '.join(pending_members)}"
        else:
            message += "🎉 Todos os membros estão com os pagamentos em dia!"

        message += price_message

        # Envia a mensagem
        await channel.send(message)


@bot.command(name='setdatapgamento', help='Define o dia de pagamento do mês')
async def set_payment_day(ctx, day: int):
    """Define o novo dia de pagamento e salva a alteração"""
    if not ctx.author.guild_permissions.administrator:
        await ctx.send("❌ Você não tem permissão para alterar o dia de pagamento.")
        return
    if day < 1 or day > 31:
        await ctx.send("❌ O dia de pagamento deve ser entre 1 e 31.")
        return

    # Atualiza o dia de pagamento e salva
    payment_manager.set_payment_day(day)
    payment_manager.upload_to_google_drive()
    # Confirmação para o usuário
    await ctx.send(f"✅ O dia de pagamento foi atualizado para o dia {day} do mês.")


@bot.command(name='historico', help='Exibe o histórico de pagamentos de um usuário específico')
async def payment_history(ctx, member: discord.Member = None):
    """Exibe o histórico de pagamentos de um membro específico"""
    if member is None:
        await ctx.send("❌ Você precisa mencionar um membro. Exemplo: `!historico @usuario`")
        return

    user_id = str(member.id)
    if user_id not in payment_manager.data["members"]:
        await ctx.send(f"❌ Nenhum histórico encontrado para {member.name}.")
        return

    payments = payment_manager.data["members"][user_id]["payments"]
    if not payments:
        await ctx.send(f"📋 {member.name} não tem nenhum pagamento registrado.")
        return

    history = "\n".join(
        [f"📅 {month}: {'✅ Pago' if paid else '❌ Não Pago'}" for month, paid in sorted(payments.items())]
    )

    await ctx.send(f"📊 **Histórico de pagamentos de {member.name}:**\n{history}")

@bot.command(name='linkcontas', help='Vincula uma conta secundária a uma conta principal.')
async def link_accounts(ctx, main_user: discord.Member, secondary_user: discord.Member):
    """Vincula uma conta secundária a uma principal"""
    main_id = str(main_user.id)
    secondary_id = str(secondary_user.id)

    if secondary_id in payment_manager.data["account_links"]:
        await ctx.send("❌ A conta secundária já está vinculada a outra principal.")
        return

    payment_manager.link_account(secondary_id, main_id)
    payment_manager.upload_to_google_drive()
    await ctx.send(f"✅ Conta {secondary_user.mention} foi vinculada à principal {main_user.mention}.")

@bot.command(name='addmembro', help='Adiciona um novo membro ao sistema de pagamentos')
async def add_member(ctx, member: discord.Member):
    """Adiciona um novo membro ao sistema de pagamentos."""
    user_id = str(member.id)
    username = member.name

    # Verifica se o membro já está registrado
    if user_id in payment_manager.data["members"]:
        await ctx.send(f"❌ O membro {member.mention} já está registrado no sistema.")
        return

    # Adiciona o membro ao sistema com o status inicial de pagamento pendente
    payment_manager.register_initial_payment(user_id, username)
    payment_manager.upload_to_google_drive()

    await ctx.send(f"✅ O membro {member.mention} foi adicionado ao sistema com o status de pagamento pendente.")

@bot.command(name='removemembro', help='Remove um membro do sistema de pagamentos')
async def remove_member(ctx, member: discord.Member):
    """Remove um membro do sistema de pagamentos."""
    user_id = str(member.id)

    # Verifica se o membro está registrado
    if user_id not in payment_manager.data["members"]:
        await ctx.send(f"❌ O membro {member.mention} não está registrado no sistema.")
        return

    # Remove o membro do sistema
    del payment_manager.data["members"][user_id]

    # Remove links relacionados a este membro (se houver)
    payment_manager.data["account_links"] = {
        sec_id: main_id for sec_id, main_id in payment_manager.data["account_links"].items() if main_id != user_id
    }

    # Salva as alterações no arquivo Excel
    payment_manager.save_data()
    payment_manager.upload_to_google_drive()

    await ctx.send(f"✅ O membro {member.mention} foi removido do sistema de pagamentos.")

@bot.command(name='limpar', help='Limpa todas as mensagens do canal atual.')
@commands.has_permissions(manage_messages=True)
async def clear_channel(ctx):
    """Limpa todas as mensagens do canal atual."""
    try:
        await ctx.channel.purge(limit=None)  # Exclui todas as mensagens do canal
        await ctx.send("✅ Todas as mensagens foram apagadas com sucesso.", delete_after=5)
    except discord.Forbidden:
        await ctx.send("❌ Eu não tenho permissão para apagar mensagens neste canal.")
    except discord.HTTPException as e:
        await ctx.send(f"❌ Ocorreu um erro ao tentar apagar as mensagens: {e}")

@bot.command(name='pagamentoauto', help='Configura um membro para ter pagamento automático todos os meses.')
@commands.has_permissions(administrator=True)
async def set_auto_payment(ctx, member: discord.Member):
    """Adiciona um membro à lista de pagamento automático."""
    user_id = str(member.id)
    
    if user_id in payment_manager.data["auto_paid_members"]:
        await ctx.send(f"❌ O membro {member.mention} já está configurado para pagamento automático.")
        return
    
    payment_manager.data["auto_paid_members"].append(user_id)
    payment_manager.save_data()
    payment_manager.upload_to_google_drive()
    await ctx.send(f"✅ O membro {member.mention} foi configurado para pagamento automático todos os meses.")

@bot.command(name='arquivos', help='Envia o link da pasta do Google Drive onde os arquivos estão sendo salvos')
async def send_drive_folder_link(ctx):
    """Envia o link da pasta no Google Drive onde os arquivos estão sendo salvos"""
    folder_link = f"https://drive.google.com/drive/folders/xxxxxxxxxxxxxx?usp=sharing"
    await ctx.send(f"📂 Aqui está o link para a pasta no Google Drive: {folder_link}")


@check_payments.before_loop
async def before_check_payments():
    await bot.wait_until_ready()

if __name__ == "__main__":
    bot.run(TOKEN)
